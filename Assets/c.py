import openpyxl

# Load your Excel workbook
file_path = "code_mapping_with_company_ids.xlsx"  # Replace with your actual file
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Header row detection (assuming row 1 is the header)
print("Codes (column E) where 'name' is empty and 'code_mapping_ids/company_id' is 'TBG':")

for row in range(2, ws.max_row + 1):  # Start after header
    name = ws[f"C{row}"].value       # Column C = 'name'
    code = ws[f"E{row}"].value       # Column E = 'code'
    company_id = ws[f"J{row}"].value # Column J = 'code_mapping_ids/company_id'

    if (company_id == "TBG") and (name is None or str(name).strip() == ""):
        print(f"- {code}  (row {row})")
